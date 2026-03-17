#!/usr/bin/env python3
"""
Extrai produtos da loja Nathy Pratasefolheados (OffStore) e exporta para Excel.
Usa Playwright para obter a sessão (API retorna 401 sem sessão) e captura
os produtos via API.
Uso: python extrair_produtos_offstore.py
Requer: pip install playwright openpyxl
        python -m playwright install chromium
"""

import sys
from pathlib import Path
from datetime import datetime

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Instale: pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instale: pip install openpyxl")
    sys.exit(1)


USERNAME = "nathypratasefolheados"
SHOP_URL = f"https://{USERNAME}.offstore.me/shop"
API_BASE = f"https://api.offstore.me/linkswithpaginationwithparams/{USERNAME}"
SHOP_BASE = f"https://{USERNAME}.offstore.me"
OUTPUT_FILE = Path(__file__).parent.parent / "produtos_nathy_pratas.xlsx"

TAGS = {
    365443: "Todos",
    472783: "01 JOIAS FEMININA",
    886353: "02 JOIAS MASCULINA",
}


def buscar_produtos_com_playwright(tag_id=365443, page_size=100):
    """Abre a página da loja (estabelece sessão) e faz requisições à API pelo mesmo contexto."""
    todos = []
    page_num = 1

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )

        try:
            page = context.new_page()
            print("  -> Carregando página da loja (para obter sessão)...")
            page.goto(SHOP_URL, wait_until="networkidle", timeout=60000)
            page.wait_for_timeout(2000)

            total = None
            while True:
                url = f"{API_BASE}?page={page_num}&pageSize={page_size}&tagId={tag_id}&stringSearch="
                print(f"  -> Buscando página {page_num}...", end=" ")
                try:
                    resp = context.request.get(url, timeout=30000)
                    if resp.status >= 400:
                        raise Exception(f"HTTP {resp.status}")
                    data = resp.json()
                except Exception as e:
                    print(f"Erro: {e}")
                    break

                rows = data.get("rows") or []
                total = data.get("total", 0)
                print(f"{len(rows)} produtos (total: {total})")

                for row in rows:
                    row["_categoria_api"] = TAGS.get(tag_id, "")
                    todos.append(row)

                if not rows or len(todos) >= total:
                    break
                page_num += 1

        finally:
            browser.close()

    return todos


def buscar_categorias_por_produto(context, tag_ids=None):
    """Monta mapa id -> categorias usando o contexto já autenticado."""
    tag_ids = tag_ids or [472783, 886353]  # JOIAS FEMININA, JOIAS MASCULINA
    mapa = {}
    for tag_id in tag_ids:
        nome_cat = TAGS.get(tag_id, "")
        page_num = 1
        while True:
            url = f"{API_BASE}?page={page_num}&pageSize=100&tagId={tag_id}&stringSearch="
            try:
                resp = context.request.get(url, timeout=30000)
                if resp.status != 200:
                    break
                data = resp.json()
            except Exception:
                break
            rows = data.get("rows") or []
            for row in rows:
                pid = row.get("id")
                if pid:
                    mapa.setdefault(pid, []).append(nome_cat)
            if not rows or len(rows) < 100:
                break
            page_num += 1
    return mapa


def _formatar_preco(val):
    if val is None:
        return ""
    try:
        v = float(str(val).replace(",", "."))
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return str(val)


def normalizar_produto(row, categorias_map=None):
    images = row.get("images") or []
    img_url = ""
    if images:
        first = images[0]
        if isinstance(first, dict):
            img_url = first.get("image_url") or ""
            if not img_url:
                url_rel = first.get("url", "")
                if url_rel:
                    img_url = f"https://cdn.offstoreimages.me/compressed/{url_rel}"
        elif isinstance(first, str) and first.startswith("http"):
            img_url = first

    slug = row.get("slug") or ""
    url_produto = f"{SHOP_BASE}/product/{slug}" if slug else ""

    cats = (categorias_map or {}).get(row.get("id"), [])
    categoria = ", ".join(c for c in cats if c) if cats else ""

    preco = _formatar_preco(row.get("price") or row.get("price_store"))

    return {
        "nome": (row.get("title") or "").strip(),
        "descricao": (row.get("description") or "").strip(),
        "url_foto": img_url,
        "categoria": categoria,
        "url_produto": url_produto,
        "preco": preco,
    }


def main():
    print("Abrindo navegador (modo invisível)...")
    print("Buscando produtos na OffStore...")

    produtos_raw = buscar_produtos_com_playwright(tag_id=365443)
    print(f"\n  -> Total encontrados: {len(produtos_raw)}")

    if not produtos_raw:
        print("\nNenhum produto encontrado. Verifique se a loja está acessível.")
        sys.exit(1)

    # Buscar categorias (reabre contexto rápido)
    print("  -> Buscando categorias...")
    cat_map = {}
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            )
            page = context.new_page()
            page.goto(SHOP_URL, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(1500)
            cat_map = buscar_categorias_por_produto(context)
            browser.close()
    except Exception as e:
        print(f"  -> Aviso (categorias): {e}")

    produtos = [normalizar_produto(p, cat_map) for p in produtos_raw]
    produtos = [p for p in produtos if p["nome"]]

    vistos = set()
    unicos = []
    for p, r in zip(produtos, produtos_raw):
        pid = r.get("id")
        if pid and pid not in vistos:
            vistos.add(pid)
            unicos.append(p)

    print(f"\nExportando {len(unicos)} produtos para Excel...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    headers = ["Nome", "Descrição", "URL Foto", "Categoria", "URL Produto", "Preço"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row, p in enumerate(unicos, 2):
        ws.cell(row=row, column=1, value=p["nome"])
        ws.cell(row=row, column=2, value=p["descricao"])
        ws.cell(row=row, column=3, value=p["url_foto"])
        ws.cell(row=row, column=4, value=p["categoria"])
        ws.cell(row=row, column=5, value=p["url_produto"])
        ws.cell(row=row, column=6, value=p["preco"])

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 35

    output = OUTPUT_FILE
    try:
        wb.save(output)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = Path(__file__).parent.parent / f"produtos_nathy_pratas_{ts}.xlsx"
        wb.save(output)
        print(f"\nArquivo original estava aberto. Salvo em: {output}")
    else:
        print(f"\nPlanilha salva em: {output}")
    print(f"Total de produtos exportados: {len(unicos)}")


if __name__ == "__main__":
    main()
