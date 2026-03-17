#!/usr/bin/env python3
"""
Extrai TODOS os produtos da Nathy Pratasefolheados (OffStore) e baixa todas as fotos.
- Busca produtos via API (Playwright para sessão)
- Visita cada página de produto para extrair a galeria de imagens
- Baixa as imagens em pastas por produto
- Exporta Excel com: nome, descrição, preço, categoria, url, imagens, etc.

Uso: python extrair_produtos_offstore_multifotos.py
Requer: pip install playwright openpyxl
        python -m playwright install chromium
"""

import os
import re
import json
import hashlib
import sys
import time
from pathlib import Path
from datetime import datetime
from urllib.parse import urljoin

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Instale: pip install playwright && python -m playwright install chromium")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instale: pip install openpyxl")
    sys.exit(1)


BASE_SITE = "https://nathypratasefolheados.offstore.me"
USERNAME = "nathypratasefolheados"
SHOP_URL = f"{BASE_SITE}/shop"
API_BASE = f"https://api.offstore.me/linkswithpaginationwithparams/{USERNAME}"

# Saída (ajuste conforme necessário)
OUT_DIR = Path(__file__).parent.parent / "export_offstore_multifotos"
HEADLESS = "--headless" in sys.argv  # padrão: mostra o navegador
IMG_DIR = OUT_DIR / "imagens_por_produto"
XLSX_OUT = OUT_DIR / "produtos_nathy_multifotos.xlsx"
DEBUG_DIR = OUT_DIR / "debug"

IMG_EXT_RE = re.compile(r"\.(jpg|jpeg|png|webp|gif)(\?|$)", re.IGNORECASE)

TAGS = {
    365443: "Todos",
    472783: "01 JOIAS FEMININA",
    886353: "02 JOIAS MASCULINA",
}


def safe_name(s):
    s = re.sub(r"[\\/:*?\"<>|]+", "", str(s or ""))
    s = re.sub(r"\s+", " ", s).strip()
    return s[:120] if s else "produto"


def sha1_short(s):
    return hashlib.sha1((s or "").encode("utf-8", errors="ignore")).hexdigest()[:10]


def absolutize(u, base=BASE_SITE):
    if not u:
        return ""
    u = str(u).strip()
    if u.startswith("data:") or u.startswith("blob:"):
        return ""
    if u.startswith("http://") or u.startswith("https://"):
        return u
    return urljoin(base, u)


def buscar_produtos_via_intercept(page):
    """
    Intercepta respostas da API quando a página carrega produtos.
    Clica em 'Mostrar mais' até ter todos.
    """
    produtos = []
    produtos_ids = set()

    def on_response(response):
        url = response.url
        if "linkswithpaginationwithparams" not in url:
            return
        try:
            body = response.json()
            rows = body.get("rows") or []
            for row in rows:
                pid = row.get("id")
                if pid and pid not in produtos_ids:
                    produtos_ids.add(pid)
                    produtos.append(row)
        except Exception:
            pass

    page.on("response", on_response)
    page.goto(SHOP_URL, wait_until="networkidle", timeout=60000)
    time.sleep(3)

    # Clicar "Mostrar mais" até carregar todos
    max_clicks = 50  # 332 / ~20 por página
    for click_num in range(max_clicks):
        btn = page.get_by_text("Mostrar mais").first
        try:
            btn.scroll_into_view_if_needed(timeout=3000)
            if not btn.is_visible(timeout=2000):
                break
            btn.click()
            time.sleep(2.5)
            if (click_num + 1) % 5 == 0:
                print(f"      -> Carregados {len(produtos)} produtos...")
        except Exception:
            break

    time.sleep(2)  # última resposta
    return produtos


def extract_images_from_page(page):
    """Extrai URLs de imagens da página do produto (galeria, JSON-LD, etc)."""
    urls = set()

    # 1) img tags
    imgs = page.query_selector_all("img")
    for im in imgs:
        for attr in ["src", "data-src", "data-lazy", "data-original"]:
            try:
                v = im.get_attribute(attr)
                if v and ("http" in v or v.startswith("/")):
                    v = absolutize(v)
                    if v and (IMG_EXT_RE.search(v) or "cdn" in v.lower() or "cloudinary" in v.lower()):
                        urls.add(v)
            except Exception:
                pass
        try:
            srcset = im.get_attribute("srcset")
            if srcset:
                parts = [p.strip().split()[0] for p in srcset.split(",") if p.strip()]
                if parts:
                    urls.add(absolutize(parts[-1]))
        except Exception:
            pass

    # 2) background-image
    els = page.query_selector_all("[style*='background-image']")
    for el in els:
        try:
            st = el.get_attribute("style") or ""
            m = re.search(r"background-image:\s*url\([\"']?(.*?)[\"']?\)", st)
            if m:
                v = absolutize(m.group(1))
                if v:
                    urls.add(v)
        except Exception:
            pass

    # 3) JSON-LD
    scripts = page.query_selector_all("script[type='application/ld+json']")
    for sc in scripts:
        try:
            txt = sc.inner_text() or ""
            if not txt.strip():
                continue
            data = json.loads(txt)
            stack = [data]
            while stack:
                x = stack.pop()
                if isinstance(x, dict):
                    for k, v in x.items():
                        if str(k).lower() == "image":
                            if isinstance(v, list):
                                for it in v:
                                    u = absolutize(it)
                                    if u:
                                        urls.add(u)
                            elif isinstance(v, str):
                                u = absolutize(v)
                                if u:
                                    urls.add(u)
                        else:
                            stack.append(v)
                elif isinstance(x, list):
                    for it in x:
                        stack.append(it)
        except Exception:
            pass

    # Filtrar logos/ícones
    cleaned = []
    for u in urls:
        lu = u.lower()
        if any(x in lu for x in ["logo", "favicon", "sprite", "icon"]):
            continue
        cleaned.append(u)
    return sorted(set(cleaned))


def download_image(context, url, out_path_no_ext):
    """Baixa imagem usando o contexto Playwright (mesmos cookies)."""
    url = absolutize(url)
    if not url:
        return ""
    try:
        resp = context.request.get(url, timeout=60000)
        if resp.status >= 400:
            return ""
        content = resp.body()
        ct = (resp.headers.get("content-type") or "").lower()
        ext = ".jpg"
        if "png" in ct:
            ext = ".png"
        elif "webp" in ct:
            ext = ".webp"
        elif "gif" in ct:
            ext = ".gif"
        path = str(out_path_no_ext) + ext
        with open(path, "wb") as f:
            f.write(content)
        if os.path.getsize(path) < 2500:
            try:
                os.remove(path)
            except Exception:
                pass
            return ""
        return path
    except Exception:
        return ""


def _formatar_preco(val):
    if val is None:
        return ""
    try:
        v = float(str(val).replace(",", "."))
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return str(val)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    IMG_DIR.mkdir(parents=True, exist_ok=True)
    DEBUG_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("OffStore - Extração completa (API + multifotos)")
    print("=" * 60)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        page = context.new_page()

        try:
            # 1) Carregar loja e interceptar produtos (API dá 401 sem sessão)
            print("\n[1/3] Carregando loja e clicando 'Mostrar mais' para carregar todos...")
            produtos_raw = buscar_produtos_via_intercept(page)
            print(f"      -> {len(produtos_raw)} produtos encontrados")

            if not produtos_raw:
                print("Nenhum produto encontrado. Verifique se a loja está acessível.")
                sys.exit(1)

            # Categorias (links_tags.tag_id na resposta)
            cat_map = {}
            for row in produtos_raw:
                lt = row.get("links_tags") or {}
                tid = lt.get("tag_id")
                if tid and TAGS.get(tid):
                    cat_map.setdefault(row.get("id"), []).append(TAGS[tid])

            # 3) Por cada produto: visitar página, extrair imagens, baixar
            print("\n[2/3] Visitando páginas de produto e baixando imagens...")
            resultados = []
            total_imgs = 0
            multi = 0

            for i, row in enumerate(produtos_raw):
                pid = row.get("id")
                nome = (row.get("title") or "").strip()
                desc = (row.get("description") or "").strip()
                preco = _formatar_preco(row.get("price") or row.get("price_store"))
                cats = ", ".join(cat_map.get(pid, []))
                url_produto = f"{BASE_SITE}/p/{pid}" if pid else ""

                if not url_produto:
                    continue

                folder = f"{safe_name(nome)}__{sha1_short(url_produto)}"
                prod_dir = IMG_DIR / folder
                prod_dir.mkdir(parents=True, exist_ok=True)

                imgs_urls = []
                imgs_arquivos = []
                try:
                    page.goto(url_produto, wait_until="domcontentloaded", timeout=30000)
                    time.sleep(2)

                    imgs_urls = extract_images_from_page(page)

                    if i == 0 and imgs_urls:
                        (DEBUG_DIR / "debug_imgs_primeiro_produto.txt").write_text(
                            "\n".join(imgs_urls), encoding="utf-8"
                        )

                    for idx, img_url in enumerate(imgs_urls[:25], start=1):
                        pth = download_image(
                            context,
                            img_url,
                            prod_dir / f"{idx:02d}",
                        )
                        if pth:
                            imgs_arquivos.append(pth)

                    total_imgs += len(imgs_arquivos)
                    if len(imgs_arquivos) >= 2:
                        multi += 1

                except Exception as e:
                    pass  # não interrompe

                resultados.append({
                    "nome": nome,
                    "descricao": desc,
                    "preco": preco,
                    "categoria": cats,
                    "produto_url": url_produto,
                    "imagens_urls": " | ".join(imgs_urls),
                    "imagens_arquivos": " | ".join(imgs_arquivos),
                    "qtd_imagens": len(imgs_arquivos),
                    "pasta_imagens": str(prod_dir),
                })

                if (i + 1) % 20 == 0:
                    print(f"      {i+1}/{len(produtos_raw)} | imagens: {total_imgs} | 2+ fotos: {multi}")

            # 4) Salvar Excel
            print("\n[3/3] Salvando Excel...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos"

            headers = [
                "nome", "descricao", "preco", "categoria", "produto_url",
                "imagens_urls", "imagens_arquivos", "qtd_imagens", "pasta_imagens",
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)

            for row_idx, r in enumerate(resultados, 2):
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=r.get(h, ""))

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 30

            try:
                wb.save(XLSX_OUT)
            except PermissionError:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                alt = OUT_DIR / f"produtos_nathy_multifotos_{ts}.xlsx"
                wb.save(alt)
                XLSX_OUT_FINAL = alt
            else:
                XLSX_OUT_FINAL = XLSX_OUT

            print("\n" + "=" * 60)
            print("Concluído!")
            print("  Excel:", XLSX_OUT_FINAL)
            print("  Imagens:", IMG_DIR)
            print("  Produtos:", len(resultados))
            print("  Imagens baixadas:", total_imgs)
            print("  Produtos com 2+ fotos:", multi)
            print("=" * 60)

        finally:
            browser.close()


if __name__ == "__main__":
    main()
